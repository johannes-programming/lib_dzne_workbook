import math as _math
import os as _os
import tempfile as _tmp

import lib_dzne_filedata as _fd
import openpyxl as _xl
import pandas as _pd


class WorkbookData(_fd.FileData):
    _ext = '.xlsx'
    @classmethod
    def _load(cls, /, file):
        return _xl.load_workbook(file)
    def _save(self, /, file):
        self.data.save(filename=file)
    @staticmethod
    def _default():
        return _xl.Workbook()
    @classmethod
    def clone_data(cls, data):
        with _tmp.TemporaryDirectory() as directory:
            file = _os.path.join(directory, "a" + cls.ext())
            data.save(filename=file)
            return _xl.load_workbook(file)
    @classmethod
    def from_DataFrames(cls, dataFrames):
        dataFrames = dict(dataFrames)
        if len(dataFrames) == 0:
            return None
        workbook = _xl.Workbook()
        default_sheet = workbook.active
        for table, df in dataFrames.items():
            if default_sheet is None:
                workbook.create_sheet(table)
            else:
                default_sheet.title = table
                default_sheet = None
        for sheetname, dataFrame in dataFrames.items():
            cls._datasheet(
                workbook=workbook,
                dataFrame=dataFrame,
                sheetname=sheetname,
            )
        return cls(workbook)
    @classmethod
    def get_worksheet(cls, *, workbook, sheetname=None):
        if sheetname is None:
            return workbook.active
        else:
            return workbook[sheetname]
    @classmethod
    def _datasheet(cls, *, dataFrame, workbook, sheetname=None):
        ws = cls.get_worksheet(workbook=workbook, sheetname=sheetname)
        columns = list(dataFrame.columns)
        for x, column in enumerate(columns):
            ws.cell(row=1, column=x+1).value = column
            for y, v in enumerate(dataFrame[column].tolist()):
                if _pd.isna(v):
                    continue
                elif (type(v) is float) and (_math.isinf(v)):# is this really needed?
                    value = str(v)
                else:
                    value = v
                ws.cell(row=y+2, column=x+1).value = value
    def mastersheet(self, *, dataFrame, masterrow, sheetname=None):
        """Writing data from table into masterfile-template. """
        #masterrow = BASE.config.get_config()['masterfile']['keyrow']
        wb = self.data
        if sheetname is None:
            ws = wb.active
        else:
            ws = wb[sheetname]
        for colnum in range(1, ws.max_column + 1):
            currentcell = ws.cell(column=colnum, row=masterrow)
            value = currentcell.value
            if type(value) is not str:
                continue
            value = value.strip()
            if value.startswith('='):
                continue
            if value == "":
                continue
            if value not in dataFrame.columns:
                self._set_cell(cell=currentcell, value="")
                continue
            for i, newvalue in enumerate(dataFrame[value].tolist()):
                datacell = ws.cell(column=colnum, row=masterrow+i)
                self._set_cell(cell=datacell, value=None if _pd.isna(newvalue) else newvalue)
        self.data = wb
    @classmethod
    def _set_cell(cls, *, cell, value):
        """Setting value of cell. """
        if _pd.isna(value):
            value = 'N/A'
        else:
            if type(value) is float:
                if _math.isinf(value):
                    if value < 0:
                        value = '-inf'
                    else:
                        value = '+inf'
            if type(value) not in {str, int, float, bool}:
                raise TypeError(f"The value {value} is of the invalid type {type(value)}! ")
        cell.value = value
        cell.alignment = _xl.styles.Alignment()#horizontal='general')



